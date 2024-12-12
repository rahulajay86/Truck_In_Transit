import pandas as pd
import numpy as np
from openpyxl import load_workbook
import time
import openpyxl

# Start the timer
start_time = time.time()

# Define file paths in variables
input_file_path = r"D:\Neelitech\MIS\python\Input_Files\merge\2024\merged_files.xlsx"
output_file_path = r"D:\Neelitech\MIS\python\Input_Files\merge\2024\Sales Vol_Price Analysis Report.xlsx"

# Ensure the correct reader is set
pd.set_option("io.excel.xlsx.reader", "openpyxl")

# Read Excel file with explicit engine and required columns to reduce memory usage
columns_to_read = ['SaleGrp', 'SaleOrg', 'AICMA', 'Billed Quantity', 'Amount in LC']
df = pd.read_excel(input_file_path, sheet_name='Sheet1', usecols=columns_to_read, engine='openpyxl')

# Filter SaleGrp, SaleOrg, and AICMA by specific values
filtered_df = df[
    df['SaleGrp'].isin(['DOM', 'ICC', 'INS', 'MIS']) &
    df['SaleOrg'].isin([1010, 1020, 1063, 1064, 1066, 1069, 1081]) &
    df['AICMA'].str.startswith(tuple(['KIDS', 'MTB GEARED', 'MTB SS', 'ROADSTER JUNIOR', 'ROADSTER SENIOR', 'SLR GENTS', 'SLR LADIES']))
]

# Standardize AICMA names by replacing variations with a common name
filtered_df.loc[:, 'AICMA'] = filtered_df['AICMA'].str.replace(r'MTB GEARED-.*', 'MTB GEARED', regex=True)
filtered_df.loc[:, 'AICMA'] = filtered_df['AICMA'].str.replace(r'KIDS-.*', 'KIDS', regex=True)
filtered_df.loc[:, 'AICMA'] = filtered_df['AICMA'].str.replace(r'MTB SS-.*', 'MTB SS', regex=True)
filtered_df.loc[:, 'AICMA'] = filtered_df['AICMA'].str.replace(r'ROADSTER SENIOR-.*', 'ROADSTER SENIOR', regex=True)
filtered_df.loc[:, 'AICMA'] = filtered_df['AICMA'].str.replace(r'SLR GENTS-.*', 'SLR GENTS', regex=True)
filtered_df.loc[:, 'AICMA'] = filtered_df['AICMA'].str.replace(r'SLR LADIES-.*', 'SLR LADIES', regex=True)

# Ensure that both columns are cast to the correct data types before assignment
filtered_df.loc[:, 'Billed Quantity'] = filtered_df['Billed Quantity'].astype(float)
filtered_df.loc[:, 'Amount in LC'] = filtered_df['Amount in LC'].astype(float)

# Perform the swap after explicit casting
filtered_df.loc[:, ['Billed Quantity', 'Amount in LC']] = filtered_df[['Amount in LC', 'Billed Quantity']].values.astype(float)


# Create pivot table to sum Billed Quantity and Amount in LC grouped by AICMA
pivot = filtered_df.pivot_table(
    index=['AICMA'],
    values=['Billed Quantity', 'Amount in LC'],
    aggfunc='sum',
    margins=True,
    margins_name='Grand Total'
)

# Rename the columns
pivot.columns = ['Sum of Billed Quantity', 'Sum of Amount in LC']

# Load the existing workbook
book = load_workbook(input_file_path)

# Remove the existing sheet if it exists
if 'Sheet3' in book.sheetnames:
    del book['Sheet3']
    book.save(input_file_path)

# Write the pivot table to a new sheet named 'Sheet3'
with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a') as writer:
    pivot.to_excel(writer, sheet_name='Sheet3')

# Divide the 'Sum of Billed Quantity' by 1000 and 'Sum of Amount in LC' by 1 crore using numpy for accurate rounding
pivot['Sum of Billed Quantity'] = np.ceil(pivot['Sum of Billed Quantity'] / 1000 * 10) / 10
pivot['Sum of Amount in LC'] = np.ceil(pivot['Sum of Amount in LC'] / 10000000 * 10) / 10

# Define the desired order of SEG1
order = ['KIDS', 'ROADSTER SENIOR', 'ROADSTER JUNIOR', 'MTB SS', 'MTB GEARED', 'SLR GENTS', 'SLR LADIES']
pivot = pivot.reindex(order)

print(pivot)

# Open the target workbook and select the specific sheet (FTM Jul)
target_wb = load_workbook(output_file_path)
target_sheet = target_wb['FTM Jul']
# target_sheet = target_wb['FTM']

# Paste the 'Sum of Billed Quantity' and 'Sum of Amount in LC' starting from column E (5th column), row 5 to row 11
for i, row in enumerate(pivot.itertuples(), start=5):
    target_sheet.cell(row=i, column=5, value=row[1])  # 'Sum of Billed Quantity' in column E
    target_sheet.cell(row=i, column=6, value=row[2])  # 'Sum of Amount in LC' in column F

# Save the changes to the target workbook
target_wb.save(output_file_path)

# End the timer and calculate the elapsed time
end_time = time.time()
execution_time = end_time - start_time

# Convert execution time to minutes and seconds
minutes, seconds = divmod(execution_time, 60)

print(f"Pivot table successfully written to 'Sheet3'.")
print(f"Pivot table data successfully pasted into 'Sales Vol & Price Analysis Report.xlsx'.")
print(f"Execution time: {int(minutes)} minutes and {seconds:.2f} seconds")
